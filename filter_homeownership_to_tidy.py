#!/usr/bin/env python3
"""Parse Census Table 3 homeownership xlsx into a tidy CSV for a state subset."""

from __future__ import annotations

import argparse
import csv
import re
from pathlib import Path

import openpyxl

# Default paths relative to this script's directory.
SCRIPT_DIR = Path(__file__).resolve().parent
DEFAULT_SOURCE = SCRIPT_DIR / "tab3_state05_2026_hmr.xlsx"
DEFAULT_STATES = SCRIPT_DIR / "List of States.xlsx"
DEFAULT_OUTPUT = SCRIPT_DIR / "homeownership_27_states_tidy.csv"

STATE_ABBR_TO_NAME = {
    "AR": "Arkansas",
    "AZ": "Arizona",
    "CA": "California",
    "CO": "Colorado",
    "CT": "Connecticut",
    "DC": "District of Columbia",
    "DE": "Delaware",
    "FL": "Florida",
    "IL": "Illinois",
    "MA": "Massachusetts",
    "MD": "Maryland",
    "ME": "Maine",
    "MN": "Minnesota",
    "NH": "New Hampshire",
    "NJ": "New Jersey",
    "NM": "New Mexico",
    "NY": "New York",
    "OH": "Ohio",
    "OR": "Oregon",
    "PA": "Pennsylvania",
    "RI": "Rhode Island",
    "TX": "Texas",
    "UT": "Utah",
    "VA": "Virginia",
    "VT": "Vermont",
    "WA": "Washington",
    "WI": "Wisconsin",
}

QUARTER_LABELS = ("Q1", "Q2", "Q3", "Q4")
# Value / margin-of-error pairs start at column index 1 (0-based).
RATE_COLS = (1, 3, 5, 7)
MOE_COLS = (2, 4, 6, 8)


def clean_state_name(raw: str) -> str:
    """Strip Census dot leaders from state names."""
    return raw.split(".")[0].strip()


def normalize_name(name: str) -> str:
    return re.sub(r"\s+", " ", name.strip().lower())


def load_state_filter(states_path: Path) -> tuple[set[str], dict[str, str]]:
    """Return allowed abbreviations and abbr -> canonical full name."""
    wb = openpyxl.load_workbook(states_path, read_only=True, data_only=True)
    ws = wb.active
    abbrs: set[str] = set()
    for row in ws.iter_rows(values_only=True):
        if not row or not row[0]:
            continue
        abbr = str(row[0]).strip().upper()
        if abbr in STATE_ABBR_TO_NAME:
            abbrs.add(abbr)
    wb.close()

    name_to_abbr = {
        normalize_name(full): abbr for abbr, full in STATE_ABBR_TO_NAME.items() if abbr in abbrs
    }
    return abbrs, name_to_abbr


def is_state_row(cell) -> bool:
    return isinstance(cell, str) and "...." in cell


def parse_state_row(row: tuple, year: int, name_to_abbr: dict[str, str]) -> list[dict]:
    """Extract up to four quarterly records from one state row."""
    state_name = clean_state_name(row[0])
    norm = normalize_name(state_name)
    abbr = name_to_abbr.get(norm)
    if abbr is None:
        return []

    records: list[dict] = []
    row = list(row)
    for q_label, rate_idx, moe_idx in zip(QUARTER_LABELS, RATE_COLS, MOE_COLS):
        rate = row[rate_idx] if rate_idx < len(row) else None
        if rate is None:
            continue
        moe = row[moe_idx] if moe_idx < len(row) else None
        records.append(
            {
                "state_abbr": abbr,
                "state_name": STATE_ABBR_TO_NAME[abbr],
                "year": year,
                "quarter": q_label,
                "homeownership_rate_pct": float(rate),
                "margin_of_error": float(moe) if moe is not None else "",
            }
        )
    return records


def parse_modern_blocks(ws, name_to_abbr: dict[str, str]) -> list[dict]:
    """Parse 2015-2026 blocks: year embedded in quarter column headers."""
    records: list[dict] = []
    rows = list(ws.iter_rows(values_only=True))
    i = 0
    while i < len(rows):
        row = rows[i]
        if not row or row[0] != "State" or not row[1] or "Quarter" not in str(row[1]):
            i += 1
            continue

        year_match = re.search(r"(20\d{2})", str(row[1]))
        if not year_match:
            i += 1
            continue
        year = int(year_match.group(1))
        i += 1

        while i < len(rows):
            row = rows[i]
            if not row or not row[0]:
                i += 1
                continue
            if isinstance(row[0], str) and row[0].startswith("Table 3"):
                break
            if is_state_row(row[0]):
                records.extend(parse_state_row(row, year, name_to_abbr))
            i += 1
    return records


def parse_legacy_blocks(ws, name_to_abbr: dict[str, str]) -> list[dict]:
    """Parse 2006-2014 blocks: year repeated in a sub-header row."""
    records: list[dict] = []
    rows = list(ws.iter_rows(values_only=True))
    i = 0
    while i < len(rows):
        row = rows[i]
        # Year row: [None, 2014, None, 2014, None, 2014, None, 2014, ...]
        if row and row[0] is None and isinstance(row[1], int) and 2000 <= row[1] <= 2030:
            year = int(row[1])
            if not (
                len(row) > 7
                and row[3] == year
                and row[5] == year
                and row[7] == year
            ):
                i += 1
                continue
            i += 1
            while i < len(rows):
                row = rows[i]
                if not row or not row[0]:
                    i += 1
                    continue
                if isinstance(row[0], str) and row[0].startswith("Table 3"):
                    break
                if row[0] == "State":
                    break
                if is_state_row(row[0]):
                    records.extend(parse_state_row(row, year, name_to_abbr))
                i += 1
            continue
        i += 1
    return records


def dedupe_records(records: list[dict]) -> list[dict]:
    """Keep last occurrence if the same state/year/quarter appears twice."""
    seen: dict[tuple, dict] = {}
    for rec in records:
        key = (rec["state_abbr"], rec["year"], rec["quarter"])
        seen[key] = rec
    return sorted(seen.values(), key=lambda r: (r["state_abbr"], r["year"], r["quarter"]))


def parse_homeownership(source_path: Path, states_path: Path) -> list[dict]:
    abbrs, name_to_abbr = load_state_filter(states_path)
    if not abbrs:
        raise ValueError(f"No valid state abbreviations found in {states_path}")

    wb = openpyxl.load_workbook(source_path, read_only=True, data_only=True)
    ws = wb.active
    records = parse_modern_blocks(ws, name_to_abbr)
    records.extend(parse_legacy_blocks(ws, name_to_abbr))
    wb.close()

    records = dedupe_records(records)
    records = [r for r in records if r["state_abbr"] in abbrs]
    return records


def write_csv(records: list[dict], output_path: Path) -> None:
    fieldnames = [
        "state_abbr",
        "state_name",
        "year",
        "quarter",
        "homeownership_rate_pct",
        "margin_of_error",
    ]
    with output_path.open("w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(records)


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Filter Census state homeownership data to a tidy CSV."
    )
    parser.add_argument("--source", type=Path, default=DEFAULT_SOURCE)
    parser.add_argument("--states", type=Path, default=DEFAULT_STATES)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    args = parser.parse_args()

    records = parse_homeownership(args.source, args.states)
    write_csv(records, args.output)

    states = sorted({r["state_abbr"] for r in records})
    years = sorted({r["year"] for r in records})
    print(f"Wrote {len(records)} rows -> {args.output}")
    print(f"States: {len(states)} ({', '.join(states)})")
    print(f"Years: {years[0]}-{years[-1]} ({len(years)} years)")


if __name__ == "__main__":
    main()
