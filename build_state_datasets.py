#!/usr/bin/env python3
"""
build_state_datasets.py
─────────────────────────────────────────────────────────────────────────────
Build all state-level economic datasets for the 27 target states.

Outputs (in ./output):
  • real_gdp_per_capita_by_state.csv
  • personal_income_per_capita_by_state.csv
  • monthly_unemployment_rate_by_state.csv
  • unemployment_rate_by_state.csv
  • homeownership_rate_by_state.csv
  • source_notes.csv
  • data_quality_report.csv

Professor-ready cleaned copies:
  • ./cleaned_output/                      — 2010Q1–2025Q4 panel (unchanged)
  • ./cleaned_output_2001_2025/            — 2001–2025 extended panel (partial GDP/home)
  • ./cleaned_output_2001_2025_full_panel/ — 2001–2025 full panel (annual backfill)

HTTP downloads use Scrapling (curl_cffi browser impersonation) to avoid FRED
Akamai blocks on plain Python requests.

Run:
  python3 build_state_datasets.py              # 2010 + extended 2001 panels
  python3 build_state_datasets.py --extended   # 2001–2025 panel only
  python3 build_state_datasets.py --full-panel # 2001–2025 full panel (RGSP + Table 13)
─────────────────────────────────────────────────────────────────────────────
"""

from __future__ import annotations

import argparse
import calendar
from dataclasses import dataclass
import logging
import os
import random
import re
import sys
import time
from io import BytesIO, StringIO
from pathlib import Path

import pandas as pd

# Quiet Scrapling per-request INFO logs (they look like the script is stuck).
logging.basicConfig(level=logging.WARNING)
for _logger in ("scrapling", "scrapling.core.utils"):
    logging.getLogger(_logger).setLevel(logging.WARNING)

# Scrapling — curl_cffi-based fetcher that bypasses FRED TLS fingerprint blocks
SCRAPLING_ROOT = Path("/Users/samsonbui/Documents/Github-2026/Scrapling")
if str(SCRAPLING_ROOT) not in sys.path:
    sys.path.insert(0, str(SCRAPLING_ROOT))

from scrapling.fetchers import FetcherSession  # noqa: E402

# ── Configuration ─────────────────────────────────────────────────────────────
SCRIPT_DIR = Path(__file__).resolve().parent
OUTPUT_DIR = SCRIPT_DIR / "output"
OUTPUT_DIR.mkdir(exist_ok=True)
CLEANED_OUTPUT_DIR = SCRIPT_DIR / "cleaned_output"
EXTENDED_OUTPUT_DIR = SCRIPT_DIR / "output_2001_2025"
EXTENDED_CLEANED_DIR = SCRIPT_DIR / "cleaned_output_2001_2025"
FULL_PANEL_OUTPUT_DIR = SCRIPT_DIR / "output_2001_2025_full_panel"
FULL_PANEL_CLEANED_DIR = SCRIPT_DIR / "cleaned_output_2001_2025_full_panel"

END_DATE = "2025-12-31"
ANNUAL_BACKFILL_END_YEAR = 2004
QUARTERLY_MERGE_START_YEAR = 2005
CENSUS_TABLE13_URL = (
    "https://www.census.gov/housing/hvs/files/annual{yy:02d}/ann{yy:02d}t13.txt"
)
CENSUS_ANNUAL_CACHE_DIR = SCRIPT_DIR / "census_hvs_annual_cache"

# Legacy aliases used by the 2010 panel (kept for backward compatibility).
START_DATE = "2010-01-01"
PANEL_START_YEAR = 2010
PANEL_END_YEAR = 2025
PANEL_END_QUARTER = 4
PANEL_END_MONTH = 12


@dataclass(frozen=True)
class PanelConfig:
    label: str
    start_date: str
    end_date: str
    panel_start_year: int
    panel_end_year: int
    panel_end_quarter: int
    panel_end_month: int
    output_dir: Path
    cleaned_dir: Path
    fred_cache_dir: Path
    quarterly_expected_rows: int
    monthly_expected_rows: int
    annual_backfill: bool = False


PANEL_2010 = PanelConfig(
    label="2010-2025",
    start_date="2010-01-01",
    end_date=END_DATE,
    panel_start_year=2010,
    panel_end_year=2025,
    panel_end_quarter=4,
    panel_end_month=12,
    output_dir=OUTPUT_DIR,
    cleaned_dir=CLEANED_OUTPUT_DIR,
    fred_cache_dir=OUTPUT_DIR / ".fred_cache",
    quarterly_expected_rows=1728,
    monthly_expected_rows=5184,
)

PANEL_2001 = PanelConfig(
    label="2001-2025",
    start_date="2001-01-01",
    end_date=END_DATE,
    panel_start_year=2001,
    panel_end_year=2025,
    panel_end_quarter=4,
    panel_end_month=12,
    output_dir=EXTENDED_OUTPUT_DIR,
    cleaned_dir=EXTENDED_CLEANED_DIR,
    fred_cache_dir=EXTENDED_OUTPUT_DIR / ".fred_cache",
    quarterly_expected_rows=2700,
    monthly_expected_rows=8100,
    annual_backfill=False,
)

PANEL_2001_FULL = PanelConfig(
    label="2001-2025 full panel",
    start_date="2001-01-01",
    end_date=END_DATE,
    panel_start_year=2001,
    panel_end_year=2025,
    panel_end_quarter=4,
    panel_end_month=12,
    output_dir=FULL_PANEL_OUTPUT_DIR,
    cleaned_dir=FULL_PANEL_CLEANED_DIR,
    fred_cache_dir=EXTENDED_OUTPUT_DIR / ".fred_cache",
    quarterly_expected_rows=2700,
    monthly_expected_rows=8100,
    annual_backfill=True,
)

HOMEOWNERSHIP_FILE = SCRIPT_DIR / "tab3_state05_2026_hmr.xlsx"
STATES_FILE = SCRIPT_DIR / "List of States.xlsx"

FRED_API_KEY = os.environ.get("FRED_API_KEY", None)
FRED_CACHE_DIR = OUTPUT_DIR / ".fred_cache"
FRED_CACHE_DIR.mkdir(exist_ok=True)

BLS_LANRDERR_URL = "https://www.bls.gov/web/laus/lanrderr.xlsx"
BLS_LANRDERR_CACHE = FRED_CACHE_DIR / "lanrderr.xlsx"

# Only AR and IL publish quarterly {ABBR}OPCI on FRED; others use OTOT/POP.
OPCI_STATES = frozenset({"AR", "IL"})

# Per-request pacing (seconds). Set FRED_PAUSE=0 to disable.
FRED_PAUSE = float(os.environ.get("FRED_PAUSE", "0.05"))

TARGET_ABBRS = frozenset({
    "AR", "AZ", "CA", "CO", "CT", "DC", "DE", "FL", "IL", "MA", "MD", "ME",
    "MN", "NH", "NJ", "NM", "NY", "OH", "OR", "PA", "RI", "TX", "UT", "VA",
    "VT", "WA", "WI",
})

STATE_MAP = {
    "AR": "Arkansas",
    "AZ": "Arizona",
    "CA": "California",
    "CO": "Colorado",
    "CT": "Connecticut",
    "DC": "District of Columbia",
    "DE": "Delaware",
    "FL": "Florida",
    "IL": "Illinois",
    "MA": "Massachusetts",
    "MD": "Maryland",
    "ME": "Maine",
    "MN": "Minnesota",
    "NH": "New Hampshire",
    "NJ": "New Jersey",
    "NM": "New Mexico",
    "NY": "New York",
    "OH": "Ohio",
    "OR": "Oregon",
    "PA": "Pennsylvania",
    "RI": "Rhode Island",
    "TX": "Texas",
    "UT": "Utah",
    "VA": "Virginia",
    "VT": "Vermont",
    "WA": "Washington",
    "WI": "Wisconsin",
}

QTR_END_MONTH = {1: 3, 2: 6, 3: 9, 4: 12}

# Track failed FRED series across sections
FAILED_SERIES: list[str] = []


def log(msg: str) -> None:
    print(msg, flush=True)


def _fred_cache_path(series_id: str, cache_dir: Path | None = None) -> Path:
    base = cache_dir or FRED_CACHE_DIR
    return base / f"{series_id}.csv"


def _panel_target_start_date(panel: PanelConfig, is_monthly: bool) -> str:
    if is_monthly:
        return f"{panel.panel_start_year}-01-31"
    month = QTR_END_MONTH[1]
    day = calendar.monthrange(panel.panel_start_year, month)[1]
    return f"{panel.panel_start_year}-{month:02d}-{day:02d}"


def filter_quarterly_panel(df: pd.DataFrame, panel: PanelConfig) -> pd.DataFrame:
    """Keep panel.panel_start_year Q1 through panel.panel_end_year Q4 only."""
    mask = (
        (df["year"] > panel.panel_start_year)
        | ((df["year"] == panel.panel_start_year) & (df["quarter"] >= 1))
    ) & (
        (df["year"] < panel.panel_end_year)
        | ((df["year"] == panel.panel_end_year) & (df["quarter"] <= panel.panel_end_quarter))
    )
    return df.loc[mask].copy()


def filter_monthly_panel(df: pd.DataFrame, panel: PanelConfig) -> pd.DataFrame:
    """Keep panel start month through panel end month only (excludes 2026)."""
    mask = (
        (df["year"] > panel.panel_start_year)
        | ((df["year"] == panel.panel_start_year) & (df["month"] >= 1))
    ) & (
        (df["year"] < panel.panel_end_year)
        | ((df["year"] == panel.panel_end_year) & (df["month"] <= panel.panel_end_month))
    )
    return df.loc[mask].copy()


def fill_interior_monthly_gaps(
    df: pd.DataFrame,
    value_col: str,
    panel: PanelConfig,
) -> pd.DataFrame:
    """
    Expand to the full monthly grid for every target state and linearly
    interpolate interior missing values (e.g. FRED/BLS blank for 2025M10).
    """
    expected: list[dict[str, object]] = []
    for abbr, state_name in STATE_MAP.items():
        for year in range(panel.panel_start_year, panel.panel_end_year + 1):
            month_end = panel.panel_end_month if year == panel.panel_end_year else 12
            for month in range(1, month_end + 1):
                expected.append({
                    "state_abbr": abbr,
                    "state": state_name,
                    "year": year,
                    "month": month,
                })
    grid = pd.DataFrame(expected)
    merged = grid.merge(
        df[["state_abbr", "year", "month", value_col]],
        on=["state_abbr", "year", "month"],
        how="left",
    )

    filled_parts: list[pd.DataFrame] = []
    fill_rows: list[tuple[str, int, int]] = []
    for abbr, grp in merged.groupby("state_abbr", sort=False):
        g = grp.sort_values(["year", "month"]).copy()
        missing = g[value_col].isna()
        g[value_col] = g[value_col].interpolate(method="linear", limit_area="inside")
        newly_filled = missing & g[value_col].notna()
        for row in g.loc[newly_filled, ["year", "month"]].itertuples(index=False):
            fill_rows.append((abbr, int(row.year), int(row.month)))
        filled_parts.append(g)

    out = pd.concat(filled_parts, ignore_index=True)
    if fill_rows:
        sample = ", ".join(f"{a} {y}M{m}" for a, y, m in fill_rows[:3])
        extra = f" (+{len(fill_rows) - 3} more)" if len(fill_rows) > 3 else ""
        log(
            f"  Filled {len(fill_rows)} missing monthly unemployment value(s) via "
            f"linear interpolation (FRED CSV blank: {sample}{extra})."
        )
    if out[value_col].isna().any():
        still_missing = out[out[value_col].isna()][["state_abbr", "year", "month"]]
        raise RuntimeError(
            "Unemployment panel still has gaps after interpolation:\n"
            f"{still_missing.head(10).to_string(index=False)}"
        )
    return out


def month_end_date(year: int, month: int) -> str:
    day = calendar.monthrange(year, month)[1]
    return f"{year}-{month:02d}-{day}"


def add_year_quarter(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    out["year"] = out["date"].dt.year
    out["quarter"] = out["date"].dt.quarter
    return out


def quarter_end_date(year: int, quarter: int) -> str:
    month = QTR_END_MONTH[quarter]
    day = calendar.monthrange(year, month)[1]
    return f"{year}-{month:02d}-{day}"


def finalize_quarterly(
    df: pd.DataFrame,
    value_col: str,
    panel: PanelConfig,
) -> pd.DataFrame:
    """Filter target states, panel range, dedupe, quarter-end dates."""
    out = df[df["state_abbr"].isin(TARGET_ABBRS)].copy()
    out = filter_quarterly_panel(out, panel)
    out["date"] = out.apply(
        lambda r: quarter_end_date(int(r["year"]), int(r["quarter"])), axis=1
    )
    out = (
        out.sort_values(["state_abbr", "year", "quarter"])
        .drop_duplicates(subset=["state_abbr", "date"], keep="last")
        .reset_index(drop=True)
    )
    if out[value_col].isna().any():
        pass  # keep blanks per requirements
    return out


def parse_ci_half_width(ci) -> float | None:
    """Parse BLS 'low – high' 90% confidence interval text to half-width (pp)."""
    if ci is None or (isinstance(ci, float) and pd.isna(ci)):
        return None
    text = str(ci).strip()
    match = re.match(r"([\d.]+)\s*[–\-]\s*([\d.]+)", text)
    if not match:
        return None
    low, high = float(match.group(1)), float(match.group(2))
    return (high - low) / 2


def fetch_bls_unemployment_rate_moe(session: FetcherSession) -> pd.DataFrame:
    """
    Latest BLS LAUS model-based 90% confidence interval half-width by state.

    Source: lanrderr.xlsx (current release month). Used as reference to scale
    margin of error across the historical panel (see attach_scaled_unemployment_moe).
    """
    import openpyxl

    if BLS_LANRDERR_CACHE.exists():
        wb = openpyxl.load_workbook(BLS_LANRDERR_CACHE, data_only=True)
    else:
        resp = session.get(BLS_LANRDERR_URL, timeout=30)
        BLS_LANRDERR_CACHE.write_bytes(resp.body)
        wb = openpyxl.load_workbook(BytesIO(resp.body), data_only=True)

    ws = wb.active
    name_to_abbr = {v: k for k, v in STATE_MAP.items()}
    rows: list[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] not in name_to_abbr:
            continue
        abbr = name_to_abbr[row[0]]
        if abbr not in TARGET_ABBRS:
            continue
        rate, ci = row[1], row[2]
        half = parse_ci_half_width(ci)
        if half is None or rate is None or float(rate) == 0:
            continue
        rows.append({"state_abbr": abbr, "ref_rate": float(rate), "ref_moe": half})
    wb.close()

    if len(rows) < len(TARGET_ABBRS):
        log(
            f"  WARNING: BLS unemployment MOE matched {len(rows)}/"
            f"{len(TARGET_ABBRS)} target states"
        )
    return pd.DataFrame(rows)


def attach_scaled_unemployment_moe(
    df: pd.DataFrame,
    rate_col: str,
    moe_ref: pd.DataFrame,
) -> pd.DataFrame:
    """Scale latest BLS 90% MOE to each row's unemployment rate (percentage points)."""
    out = df.merge(moe_ref, on="state_abbr", how="left")
    out["margin_of_error"] = out[rate_col] * (out["ref_moe"] / out["ref_rate"])
    return out.drop(columns=["ref_rate", "ref_moe"])


def finalize_monthly(
    df: pd.DataFrame,
    value_col: str,
    panel: PanelConfig,
) -> pd.DataFrame:
    out = df[df["state_abbr"].isin(TARGET_ABBRS)].copy()
    out = filter_monthly_panel(out, panel)
    out["date"] = out.apply(
        lambda r: month_end_date(int(r["year"]), int(r["month"])), axis=1
    )
    out = (
        out.sort_values(["state_abbr", "year", "month"])
        .drop_duplicates(subset=["state_abbr", "date"], keep="last")
        .reset_index(drop=True)
    )
    return out


# ── FRED fetcher ──────────────────────────────────────────────────────────────
def _parse_csv_response(resp_text: str, series_id: str, start: str) -> pd.DataFrame:
    df = pd.read_csv(StringIO(resp_text))
    if df.shape[1] < 2:
        raise ValueError(f"Unexpected FRED response for {series_id}: {df.head()}")
    df.columns = ["date", "value"]
    df["date"] = pd.to_datetime(df["date"], errors="coerce")
    df["value"] = pd.to_numeric(df["value"], errors="coerce")
    df = df.dropna(subset=["date"]).copy()
    df = df[df["date"] >= pd.Timestamp(start)].reset_index(drop=True)
    if df.empty:
        raise ValueError(f"No data returned for {series_id} after {start}")
    return df


def _fetch_fred_csv(
    session: FetcherSession,
    series_id: str,
    start: str,
    cache_dir: Path,
    timeout: int = 20,
) -> pd.DataFrame:
    cache_path = _fred_cache_path(series_id, cache_dir)
    if cache_path.exists():
        return _parse_csv_response(cache_path.read_text(encoding="utf-8"), series_id, start)

    url = (
        f"https://fred.stlouisfed.org/graph/fredgraph.csv"
        f"?id={series_id}"
        f"&cosd={start}"
    )
    resp = session.get(url, timeout=timeout, stealthy_headers=True)
    if resp.status == 404:
        raise RuntimeError(f"Series not found (404): {series_id}")
    if resp.status != 200:
        raise RuntimeError(f"HTTP {resp.status} for {series_id}")
    text = resp.body.decode("utf-8")
    if not text.strip():
        raise ValueError(f"Empty response body for {series_id}")
    cache_path.parent.mkdir(parents=True, exist_ok=True)
    cache_path.write_text(text, encoding="utf-8")
    return _parse_csv_response(text, series_id, start)


def _fetch_fred_api(series_id: str, start: str, timeout: int = 30) -> pd.DataFrame:
    if not FRED_API_KEY:
        raise RuntimeError("No FRED_API_KEY set — cannot use API fallback")
    url = (
        "https://api.stlouisfed.org/fred/series/observations"
        f"?series_id={series_id}"
        f"&observation_start={start}"
        f"&file_type=json"
        f"&api_key={FRED_API_KEY}"
    )
    with FetcherSession(impersonate="chrome", stealthy_headers=True, timeout=timeout) as session:
        resp = session.get(url)
    if resp.status != 200:
        raise RuntimeError(f"FRED API HTTP {resp.status} for {series_id}")
    obs = resp.json().get("observations", [])
    if not obs:
        raise ValueError(f"No observations from API for {series_id}")
    df = pd.DataFrame(obs)[["date", "value"]]
    df["date"] = pd.to_datetime(df["date"], errors="coerce")
    df["value"] = pd.to_numeric(df["value"], errors="coerce")
    df = df.dropna(subset=["date"]).copy()
    df = df[df["date"] >= pd.Timestamp(start)].reset_index(drop=True)
    if df.empty:
        raise ValueError(f"No data from API for {series_id} after {start}")
    return df


def fetch_fred(
    session: FetcherSession,
    series_id: str,
    start: str,
    cache_dir: Path,
    max_retries: int = 2,
) -> pd.DataFrame:
    last_error = None
    for attempt in range(1, max_retries + 1):
        try:
            return _fetch_fred_csv(session, series_id, start, cache_dir)
        except Exception as e:
            last_error = e
            if "404" in str(e) or "not found" in str(e).lower():
                break
            if attempt < max_retries:
                wait = min(2 ** attempt + random.uniform(0, 1), 8)
                log(f"     retry {attempt}/{max_retries} for {series_id}: {e} ({wait:.1f}s)")
                time.sleep(wait)

    if FRED_API_KEY:
        log(f"     CSV failed; trying FRED API for {series_id} …")
        try:
            return _fetch_fred_api(series_id, start)
        except Exception as api_err:
            print(f"     API fallback also failed for {series_id}: {api_err}")
            last_error = api_err

    raise RuntimeError(f"Failed to download {series_id} after {max_retries} attempts: {last_error}")


def fetch_state_series(
    session: FetcherSession,
    suffix: str,
    label: str,
    panel: PanelConfig,
    pause: float | None = None,
) -> tuple[pd.DataFrame, list[str]]:
    """Download {ABBR}{suffix} for every target state."""
    if pause is None:
        pause = FRED_PAUSE
    frames: list[pd.DataFrame] = []
    failed: list[str] = []
    total = len(STATE_MAP)
    log(f"\n{'=' * 65}")
    log(label)
    log(f"Series pattern : {{ABBR}}{suffix}")
    log(f"Panel          : {panel.label} (from {panel.start_date})")
    log(f"States         : {total}")
    log(f"{'=' * 65}")

    for i, (abbr, state_name) in enumerate(STATE_MAP.items(), start=1):
        series_id = f"{abbr}{suffix}"
        try:
            df = fetch_fred(session, series_id, start=panel.start_date, cache_dir=panel.fred_cache_dir)
            df["state_abbr"] = abbr
            df["state"] = state_name
            frames.append(df)
            last = df["date"].max().strftime("%Y-%m")
            cached = " (cache)" if _fred_cache_path(series_id, panel.fred_cache_dir).exists() else ""
            log(f"  [{i:2d}/{total}] ✓  {abbr:2s}  ({series_id})  {len(df)} obs  through {last}{cached}")
        except Exception as exc:
            log(f"  [{i:2d}/{total}] ✗  {abbr:2s}  ({series_id})  ERROR: {exc}")
            failed.append(series_id)
            FAILED_SERIES.append(series_id)
        if pause > 0:
            time.sleep(pause)

    if not frames:
        raise RuntimeError(f"No {suffix} data downloaded.")
    return pd.concat(frames, ignore_index=True), failed


# ── Dataset builders ──────────────────────────────────────────────────────────
def try_fetch_fred(
    session: FetcherSession,
    series_id: str,
    start: str,
    cache_dir: Path,
) -> pd.DataFrame | None:
    """Single-attempt fetch; returns None if the series does not exist."""
    try:
        return _fetch_fred_csv(session, series_id, start, cache_dir)
    except Exception as exc:
        if "404" in str(exc) or "not found" in str(exc).lower():
            return None
        raise


def population_from_fred(pop_raw: pd.DataFrame) -> pd.DataFrame:
    pop = pop_raw.copy()
    pop["year"] = pop["date"].dt.year
    pop = pop.rename(columns={"value": "pop_thousands"})
    pop["population"] = pop["pop_thousands"] * 1000
    pop_annual = (
        pop.sort_values(["state_abbr", "year", "date"])
        .groupby(["state_abbr", "year"], as_index=False)
        .agg(population=("population", "last"))
    )
    pop_annual = pop_annual.sort_values(["state_abbr", "year"])
    pop_annual["population"] = pop_annual.groupby("state_abbr")["population"].ffill()
    return pop_annual


def _personal_income_from_otot(
    session: FetcherSession,
    abbr: str,
    state_name: str,
    pop_annual: pd.DataFrame,
    panel: PanelConfig,
) -> pd.DataFrame:
    otot_id = f"{abbr}OTOT"
    otot_df = fetch_fred(session, otot_id, start=panel.start_date, cache_dir=panel.fred_cache_dir)
    otot_df["state_abbr"] = abbr
    otot_df["state"] = state_name
    otot_df = add_year_quarter(otot_df)
    otot_df = otot_df.rename(columns={"value": "personal_income_millions"})
    merged = otot_df.merge(pop_annual, on=["state_abbr", "year"], how="left")
    merged["personal_income_per_capita"] = (
        merged["personal_income_millions"] * 1_000_000 / merged["population"]
    )
    return merged


def _build_gdp_annual_assigned_quarters(
    session: FetcherSession,
    pop_annual: pd.DataFrame,
    panel: PanelConfig,
) -> pd.DataFrame:
    """
    2001–2004: annual FRED {ABBR}RGSP / {ABBR}POP, same value assigned to Q1–Q4.
    """
    rgsp_raw, rgsp_failed = fetch_state_series(
        session,
        "RGSP",
        "Annual real GDP by state (FRED {ABBR}RGSP)",
        panel,
    )
    annual = rgsp_raw.copy()
    annual["year"] = annual["date"].dt.year
    annual = annual.rename(columns={"value": "real_gdp_millions_chained_2017_dollars"})
    annual = annual[
        (annual["year"] >= panel.panel_start_year)
        & (annual["year"] <= ANNUAL_BACKFILL_END_YEAR)
    ].copy()
    annual = annual.merge(pop_annual, on=["state_abbr", "year"], how="left")
    annual["real_gdp_per_capita"] = (
        annual["real_gdp_millions_chained_2017_dollars"] * 1_000_000 / annual["population"]
    )

    rows: list[dict] = []
    for _, rec in annual.iterrows():
        for quarter in (1, 2, 3, 4):
            rows.append({
                "state": rec["state"],
                "state_abbr": rec["state_abbr"],
                "year": int(rec["year"]),
                "quarter": quarter,
                "real_gdp_millions_chained_2017_dollars": rec["real_gdp_millions_chained_2017_dollars"],
                "population": rec["population"],
                "real_gdp_per_capita": rec["real_gdp_per_capita"],
            })
    out = pd.DataFrame(rows)
    if rgsp_failed:
        log(f"  WARNING: RGSP failed series: {rgsp_failed}")
    log(
        f"  Annual RGSP assigned to quarters: {len(out)} rows "
        f"({panel.panel_start_year}–{ANNUAL_BACKFILL_END_YEAR})"
    )
    return out


def _census_state_name_to_abbr(raw_name: str) -> str | None:
    name = re.sub(r"\.+", " ", raw_name).strip()
    name = re.sub(r"\s+", " ", name)
    norm = name.lower()
    if norm.startswith("district of colu") or norm == "district of columbia":
        return "DC"
    for abbr, full in STATE_MAP.items():
        if norm == full.lower():
            return abbr
    for abbr, full in STATE_MAP.items():
        if norm.startswith(full.lower()[:12]):
            return abbr
    return None


def _parse_table13_state_line(line: str) -> tuple[str, list[float]] | None:
    line = line.strip()
    if not line or line.startswith(("Table", "/r", "United States")):
        return None
    if "Continued" in line:
        return None

    values = [float(v) for v in re.findall(r"\d+\.\d+", line)]
    if not values:
        return None
    first_num = re.search(r"\d+\.\d+", line)
    if first_num is None:
        return None
    name = re.sub(r"\.+", " ", line[: first_num.start()]).strip()
    if not name:
        return None
    return name, values


def _find_table13_year_column(lines: list[str], target_year: int) -> tuple[int, int] | None:
    """Return (header_line_index, column_index) for the last header containing target_year."""
    best: tuple[int, int] | None = None
    for i, line in enumerate(lines):
        tokens = line.split()
        year_positions = [
            (j, int(re.match(r"^(\d{4})", token).group(1)))
            for j, token in enumerate(tokens)
            if re.match(r"^\d{4}", token)
        ]
        if len(year_positions) < 2:
            continue
        for j, year in year_positions:
            if year == target_year:
                best = (i, j)
    return best


def _parse_census_table13_year(text: str, target_year: int) -> dict[str, float]:
    lines = text.splitlines()
    header = _find_table13_year_column(lines, target_year)
    if header is None:
        raise ValueError(f"Could not locate {target_year} column in Census Table 13 text")
    header_idx, col_idx = header
    rates: dict[str, float] = {}
    for line in lines[header_idx + 1:]:
        if line.strip().startswith("Table") or "Continued" in line:
            break
        parsed = _parse_table13_state_line(line)
        if not parsed:
            continue
        name, values = parsed
        if col_idx >= len(values):
            continue
        abbr = _census_state_name_to_abbr(name)
        if abbr and abbr in TARGET_ABBRS:
            rates[abbr] = values[col_idx]
    missing = sorted(TARGET_ABBRS - set(rates))
    if missing:
        raise ValueError(
            f"Census Table 13 {target_year} missing target states: {', '.join(missing)}"
        )
    return rates


def _fetch_census_table13(session: FetcherSession, year: int) -> str:
    CENSUS_ANNUAL_CACHE_DIR.mkdir(parents=True, exist_ok=True)
    cache_path = CENSUS_ANNUAL_CACHE_DIR / f"ann{year % 100:02d}t13.txt"
    if cache_path.exists():
        return cache_path.read_text(encoding="utf-8", errors="replace")

    url = CENSUS_TABLE13_URL.format(yy=year % 100)
    log(f"  Downloading Census Table 13 ({year}): {url}")
    resp = session.get(url, timeout=30, stealthy_headers=True)
    if resp.status != 200:
        raise RuntimeError(f"HTTP {resp.status} for Census Table 13 {year}")
    text = resp.body.decode("utf-8", errors="replace")
    if not text.strip():
        raise ValueError(f"Empty Census Table 13 response for {year}")
    cache_path.write_text(text, encoding="utf-8")
    return text


def _build_homeownership_annual_assigned_quarters(
    session: FetcherSession,
    years: range,
) -> pd.DataFrame:
    """Annual Census HVS Table 13 rates assigned to Q1–Q4 for each year."""
    rows: list[dict] = []
    for year in years:
        text = _fetch_census_table13(session, year)
        rates = _parse_census_table13_year(text, year)
        for abbr, rate in sorted(rates.items()):
            for quarter in (1, 2, 3, 4):
                rows.append({
                    "state": STATE_MAP[abbr],
                    "state_abbr": abbr,
                    "year": year,
                    "quarter": quarter,
                    "date": quarter_end_date(year, quarter),
                    "homeownership_rate": rate,
                    "margin_of_error": pd.NA,
                })
    out = pd.DataFrame(rows)
    log(
        f"  Census Table 13 annual → quarters: {len(out)} rows "
        f"({years.start}–{years.stop - 1})"
    )
    return out


def build_real_gdp_per_capita(
    session: FetcherSession,
    pop_annual: pd.DataFrame,
    panel: PanelConfig,
) -> pd.DataFrame:
    parts: list[pd.DataFrame] = []
    if panel.annual_backfill:
        parts.append(_build_gdp_annual_assigned_quarters(session, pop_annual, panel))

    gdp_raw, gdp_failed = fetch_state_series(
        session, "RQGSP", "Real GDP by state (FRED {ABBR}RQGSP)", panel
    )

    gdp = add_year_quarter(gdp_raw)
    gdp = gdp.rename(columns={"value": "real_gdp_millions_chained_2017_dollars"})

    merged = gdp.merge(pop_annual, on=["state_abbr", "year"], how="left")
    merged["real_gdp_per_capita"] = (
        merged["real_gdp_millions_chained_2017_dollars"] * 1_000_000 / merged["population"]
    )

    quarterly = merged[
        [
            "state",
            "state_abbr",
            "year",
            "quarter",
            "real_gdp_millions_chained_2017_dollars",
            "population",
            "real_gdp_per_capita",
        ]
    ].copy()
    if panel.annual_backfill:
        quarterly = quarterly[quarterly["year"] >= QUARTERLY_MERGE_START_YEAR].copy()
        log(
            f"  Quarterly RQGSP rows (from {QUARTERLY_MERGE_START_YEAR}): {len(quarterly)}"
        )
    parts.append(quarterly)

    combined = pd.concat(parts, ignore_index=True)
    out = finalize_quarterly(combined, "real_gdp_per_capita", panel)
    out = out[
        [
            "state",
            "state_abbr",
            "year",
            "quarter",
            "date",
            "real_gdp_per_capita",
            "real_gdp_millions_chained_2017_dollars",
            "population",
        ]
    ]
    if gdp_failed:
        log(f"  WARNING: RQGSP failed series: {gdp_failed}")
    return out


def build_personal_income_per_capita(
    session: FetcherSession,
    pop_annual: pd.DataFrame,
    panel: PanelConfig,
) -> pd.DataFrame:
    """
    Prefer FRED {ABBR}OPCI (quarterly per capita). Only AR and IL expose OPCI on
    FRED; other states fall back to {ABBR}OTOT / annual population.

    For the 2001 panel, AR and IL use OTOT/POP before OPCI begins (~2010) and
    OPCI thereafter.
    """
    frames: list[pd.DataFrame] = []
    opc_states: list[str] = []
    otot_states: list[str] = []
    hybrid_states: list[str] = []
    failed: list[str] = []

    total = len(STATE_MAP)
    log(f"\n{'=' * 65}")
    log("Personal income per capita (FRED {ABBR}OPCI or OTOT/POP)")
    log(f"Panel          : {panel.label} (from {panel.start_date})")
    log(f"States         : {total}  (OPCI direct: {', '.join(sorted(OPCI_STATES))})")
    log(f"{'=' * 65}")

    for i, (abbr, state_name) in enumerate(STATE_MAP.items(), start=1):
        if abbr in OPCI_STATES:
            opc_id = f"{abbr}OPCI"
            opc_df = try_fetch_fred(
                session, opc_id, start=panel.start_date, cache_dir=panel.fred_cache_dir
            )
            if opc_df is not None and len(opc_df) >= 4:
                if panel.panel_start_year < 2010:
                    otot_merged = _personal_income_from_otot(
                        session, abbr, state_name, pop_annual, panel
                    )
                    opc_part = opc_df.copy()
                    opc_part["state_abbr"] = abbr
                    opc_part["state"] = state_name
                    opc_part = add_year_quarter(opc_part)
                    opc_part = opc_part.rename(columns={"value": "personal_income_per_capita"})
                    opc_part = opc_part[
                        ["state", "state_abbr", "year", "quarter", "personal_income_per_capita"]
                    ]
                    combined = otot_merged[
                        ["state", "state_abbr", "year", "quarter", "personal_income_per_capita"]
                    ].merge(
                        opc_part,
                        on=["state", "state_abbr", "year", "quarter"],
                        how="left",
                        suffixes=("_otot", "_opc"),
                    )
                    combined["personal_income_per_capita"] = combined[
                        "personal_income_per_capita_opc"
                    ].fillna(combined["personal_income_per_capita_otot"])
                    frames.append(
                        combined[
                            ["state", "state_abbr", "year", "quarter", "personal_income_per_capita"]
                        ]
                    )
                    hybrid_states.append(abbr)
                    last = opc_df["date"].max().strftime("%Y-%m")
                    log(
                        f"  [{i:2d}/{total}] ✓  {abbr:2s}  ({opc_id} + OTOT/POP)  "
                        f"hybrid OPCI/OTOT  through {last}"
                    )
                else:
                    opc_df["state_abbr"] = abbr
                    opc_df["state"] = state_name
                    opc_df = add_year_quarter(opc_df)
                    opc_df = opc_df.rename(columns={"value": "personal_income_per_capita"})
                    frames.append(opc_df)
                    opc_states.append(abbr)
                    last = opc_df["date"].max().strftime("%Y-%m")
                    log(f"  [{i:2d}/{total}] ✓  {abbr:2s}  ({opc_id})  direct OPCI  through {last}")
                if FRED_PAUSE > 0:
                    time.sleep(FRED_PAUSE)
                continue

        try:
            merged = _personal_income_from_otot(session, abbr, state_name, pop_annual, panel)
            frames.append(merged)
            otot_states.append(abbr)
            last = merged["date"].max().strftime("%Y-%m")
            log(f"  [{i:2d}/{total}] ✓  {abbr:2s}  ({abbr}OTOT + POP)  computed  through {last}")
        except Exception as exc:
            log(f"  [{i:2d}/{total}] ✗  {abbr:2s}  ERROR: {exc}")
            failed.append(abbr)
            FAILED_SERIES.append(f"{abbr}OTOT")
        if FRED_PAUSE > 0:
            time.sleep(FRED_PAUSE)

    if not frames:
        raise RuntimeError("No personal income data downloaded.")

    if opc_states:
        log(f"  OPCI direct: {len(opc_states)} states ({', '.join(opc_states)})")
    if hybrid_states:
        log(f"  OPCI+OTOT hybrid: {len(hybrid_states)} states ({', '.join(hybrid_states)})")
    if otot_states:
        log(f"  OTOT/POP computed: {len(otot_states)} states")
    if failed:
        log(f"  WARNING: personal income failed: {failed}")

    raw = pd.concat(frames, ignore_index=True)
    out = raw[
        ["state", "state_abbr", "year", "quarter", "personal_income_per_capita"]
    ].copy()
    out = finalize_quarterly(out, "personal_income_per_capita", panel)
    return out[
        [
            "state",
            "state_abbr",
            "year",
            "quarter",
            "date",
            "personal_income_per_capita",
        ]
    ]


def build_unemployment(
    session: FetcherSession,
    panel: PanelConfig,
) -> tuple[pd.DataFrame, pd.DataFrame, list[str]]:
    """Monthly + quarterly unemployment; keeps complete quarters only."""
    all_monthly: list[pd.DataFrame] = []
    failed: list[str] = []

    total = len(STATE_MAP)
    log(f"\n{'=' * 65}")
    log("Unemployment rate by state (FRED BLS LAUS)")
    log("Series pattern : {ABBR}UR   e.g. ARUR, WAUR, WIUR")
    log(f"Panel          : {panel.label} (from {panel.start_date})")
    log(f"States         : {total}")
    log(f"{'=' * 65}")

    for i, (abbr, state_name) in enumerate(STATE_MAP.items(), start=1):
        series_id = f"{abbr}UR"
        try:
            df = fetch_fred(
                session, series_id, start=panel.start_date, cache_dir=panel.fred_cache_dir
            )
            if df.empty:
                raise ValueError(f"No data returned after {START_DATE}")
            df["state"] = state_name
            df["state_abbr"] = abbr
            all_monthly.append(df)
            last = df["date"].max().strftime("%Y-%m")
            log(f"  [{i:2d}/{total}] ✓  {abbr:2s}  ({series_id})  {len(df)} months  through {last}")
        except Exception as exc:
            log(f"  [{i:2d}/{total}] ✗  {abbr:2s}  ({series_id})  ERROR: {exc}")
            failed.append(abbr)
            FAILED_SERIES.append(series_id)
        if FRED_PAUSE > 0:
            time.sleep(FRED_PAUSE)

    if not all_monthly:
        raise RuntimeError("No unemployment data downloaded.")

    moe_ref = fetch_bls_unemployment_rate_moe(session)
    log(f"  BLS unemployment MOE reference: {len(moe_ref)} states (lanrderr.xlsx)")

    monthly = pd.concat(all_monthly, ignore_index=True)
    monthly["year"] = monthly["date"].dt.year
    monthly["month"] = monthly["date"].dt.month
    monthly = monthly.rename(columns={"value": "unemployment_rate"})
    monthly = monthly[monthly["state_abbr"].isin(TARGET_ABBRS)].copy()

    monthly = fill_interior_monthly_gaps(
        monthly[["state", "state_abbr", "year", "month", "unemployment_rate"]],
        "unemployment_rate",
        panel,
    )
    monthly = attach_scaled_unemployment_moe(monthly, "unemployment_rate", moe_ref)
    monthly["margin_of_error"] = monthly["margin_of_error"].round(4)
    log(
        f"  Monthly panel: {len(monthly)} rows "
        f"({monthly['state_abbr'].nunique()} states × "
        f"{len(monthly) // monthly['state_abbr'].nunique()} months)"
    )

    monthly_out = finalize_monthly(monthly, "unemployment_rate", panel)

    # Aggregate quarterly from the filled monthly panel (complete quarters only).
    panel_monthly = monthly.copy()
    panel_monthly["quarter"] = panel_monthly["month"].apply(lambda m: (m - 1) // 3 + 1)
    quarterly = (
        panel_monthly.groupby(["state", "state_abbr", "year", "quarter"], as_index=False)
        .agg(
            unemployment_rate=("unemployment_rate", "mean"),
            margin_of_error=("margin_of_error", "mean"),
        )
        .round(4)
    )
    month_counts = (
        panel_monthly.groupby(["state_abbr", "year", "quarter"])["month"]
        .count()
        .reset_index()
        .rename(columns={"month": "n_months"})
    )
    quarterly = quarterly.merge(
        month_counts, on=["state_abbr", "year", "quarter"], how="left"
    )
    quarterly = quarterly[quarterly["n_months"] == 3].drop(columns=["n_months"])

    quarterly_out = quarterly[
        ["state", "state_abbr", "year", "quarter", "unemployment_rate", "margin_of_error"]
    ].copy()
    quarterly_out = finalize_quarterly(quarterly_out, "unemployment_rate", panel)

    return monthly_out, quarterly_out, failed


def build_homeownership(
    panel: PanelConfig,
    session: FetcherSession | None = None,
) -> pd.DataFrame:
    from filter_homeownership_to_tidy import parse_homeownership

    print(f"\n{'=' * 65}")
    print("Homeownership rate by state (Census HVS)")
    print(f"Source file    : {HOMEOWNERSHIP_FILE}")
    print(f"Panel          : {panel.label}")
    print(f"{'=' * 65}")

    parts: list[pd.DataFrame] = []

    if panel.annual_backfill:
        if session is None:
            raise ValueError("FetcherSession required for annual homeownership backfill")
        annual_years = range(panel.panel_start_year, ANNUAL_BACKFILL_END_YEAR + 1)
        parts.append(_build_homeownership_annual_assigned_quarters(session, annual_years))

    if not HOMEOWNERSHIP_FILE.exists():
        raise FileNotFoundError(f"Homeownership file not found: {HOMEOWNERSHIP_FILE}")

    states_path = STATES_FILE if STATES_FILE.exists() else HOMEOWNERSHIP_FILE
    records = parse_homeownership(HOMEOWNERSHIP_FILE, states_path)
    if not records:
        raise ValueError(f"No homeownership records parsed from {HOMEOWNERSHIP_FILE}")

    home = pd.DataFrame(records)
    home = home.rename(columns={
        "state_name": "state",
        "homeownership_rate_pct": "homeownership_rate",
    })
    home["quarter"] = home["quarter"].str.replace("Q", "", regex=False).astype(int)
    home = home[home["state_abbr"].isin(TARGET_ABBRS)].copy()
    if panel.annual_backfill:
        home = home[home["year"] >= QUARTERLY_MERGE_START_YEAR].copy()
        min_q_year = int(home["year"].min()) if not home.empty else QUARTERLY_MERGE_START_YEAR
        if min_q_year > QUARTERLY_MERGE_START_YEAR:
            bridge_years = range(QUARTERLY_MERGE_START_YEAR, min_q_year)
            log(
                f"  Local Table 3 begins {min_q_year}Q1; bridging "
                f"{QUARTERLY_MERGE_START_YEAR}–{min_q_year - 1} from annual Table 13"
            )
            parts.append(
                _build_homeownership_annual_assigned_quarters(session, bridge_years)
            )
    home = filter_quarterly_panel(home, panel)
    home["date"] = home.apply(
        lambda r: quarter_end_date(int(r["year"]), int(r["quarter"])), axis=1
    )
    quarterly_out = home[
        ["state", "state_abbr", "year", "quarter", "date", "homeownership_rate", "margin_of_error"]
    ].sort_values(["state_abbr", "year", "quarter"])
    parts.append(quarterly_out)

    out = pd.concat(parts, ignore_index=True)
    out = (
        out.sort_values(["state_abbr", "year", "quarter"])
        .drop_duplicates(subset=["state_abbr", "date"], keep="last")
        .reset_index(drop=True)
    )
    print(f"  ✓  {len(out)} rows, {out['state_abbr'].nunique()} states, "
          f"{out['year'].min()}Q{out.loc[out['year']==out['year'].min(), 'quarter'].min()}"
          f" – {out['year'].max()}Q{out.loc[out['year']==out['year'].max(), 'quarter'].max()}")
    return out


def _quarterly_range_label(df: pd.DataFrame) -> str:
    ordered = df.sort_values(["year", "quarter"])
    lo = ordered.iloc[0]
    hi = ordered.iloc[-1]
    return f"{int(lo['year'])}Q{int(lo['quarter'])}-{int(hi['year'])}Q{int(hi['quarter'])}"


def _monthly_range_label(df: pd.DataFrame) -> str:
    ordered = df.sort_values(["year", "month"])
    lo = ordered.iloc[0]
    hi = ordered.iloc[-1]
    return f"{int(lo['year'])}M{int(lo['month'])}-{int(hi['year'])}M{int(hi['month'])}"


def _reaches_panel_start(
    df: pd.DataFrame,
    panel: PanelConfig,
    is_monthly: bool,
) -> bool:
    if df.empty:
        return False
    lo = df.sort_values(["year", "month" if is_monthly else "quarter"]).iloc[0]
    if is_monthly:
        return int(lo["year"]) == panel.panel_start_year and int(lo["month"]) == 1
    return int(lo["year"]) == panel.panel_start_year and int(lo["quarter"]) == 1


def build_source_notes(
    panel: PanelConfig,
    outputs: dict[str, pd.DataFrame],
) -> pd.DataFrame:
    gdp = outputs["real_gdp_per_capita_by_state.csv"]
    pi = outputs["personal_income_per_capita_by_state.csv"]
    ur_m = outputs["monthly_unemployment_rate_by_state.csv"]
    ur_q = outputs["unemployment_rate_by_state.csv"]
    home = outputs["homeownership_rate_by_state.csv"]

    gdp_range = _quarterly_range_label(gdp)
    pi_range = _quarterly_range_label(pi)
    ur_m_range = _monthly_range_label(ur_m)
    ur_q_range = _quarterly_range_label(ur_q)
    home_range = _quarterly_range_label(home)

    target_q = f"{panel.panel_start_year}Q1-{panel.panel_end_year}Q{panel.panel_end_quarter}"
    target_m = f"{panel.panel_start_year}M1-{panel.panel_end_year}M{panel.panel_end_month}"

    gdp_back = "yes" if _reaches_panel_start(gdp, panel, False) else "no"
    pi_back = "yes" if _reaches_panel_start(pi, panel, False) else "no"
    ur_back = "yes" if _reaches_panel_start(ur_m, panel, True) else "no"
    home_back = "yes" if _reaches_panel_start(home, panel, False) else "no"

    pi_method = (
        "Primary FRED {ABBR}OPCI for AR and IL from 2010 onward; OTOT/POP for 2001–2009 "
        "and for all other states across the panel."
        if panel.panel_start_year < 2010
        else "Primary FRED {ABBR}OPCI for AR and IL; OTOT/POP for other states."
    )

    if panel.annual_backfill:
        gdp_notes = (
            f"Target panel: {target_q}. Reaches {panel.panel_start_year}Q1: {gdp_back}. "
            "2001Q1–2004Q4 values are derived from annual BEA/FRED real GDP by state "
            "series {ABBR}RGSP and annual state population {ABBR}POP. Annual values are "
            "assigned to all four quarters of the corresponding year. Beginning 2005Q1, "
            "values use official quarterly BEA/FRED real GDP by state series {ABBR}RQGSP. "
            "Per-capita = real_gdp_millions_chained_2017_dollars × 1,000,000 / population; "
            "population = {ABBR}POP × 1000."
        )
        gdp_source = (
            "U.S. Bureau of Economic Analysis via FRED "
            "({ABBR}RGSP + {ABBR}POP for 2001–2004; {ABBR}RQGSP + {ABBR}POP from 2005)"
        )
        home_notes = (
            f"Target panel: {target_q}. Reaches {panel.panel_start_year}Q1: {home_back}. "
            "2001Q1–2004Q4 values are derived from Census HVS Annual Statistics, "
            "Table 13: Homeownership Rates by State. Annual values are assigned to all "
            "four quarters of the corresponding year. Beginning 2005Q1, values use Census "
            "HVS quarterly Table 3: Homeownership Rates by State, 2005-present. "
            f"Quarterly Table 3 parsed from local file {HOMEOWNERSHIP_FILE.name} "
            "(earliest quarter in file is 2006Q1); 2005Q1–2005Q4 use annual Table 13 "
            "assigned to quarters for continuity. 2026Q1 excluded. margin_of_error: "
            "Census HVS Table 3 published margin of error for quarterly rows only "
            "(percentage points, 90% confidence level); blank for annual-assigned rows."
        )
        home_source = (
            "U.S. Census Bureau HVS Annual Statistics Table 13 (2001–2004; 2005 bridge) "
            "and HVS quarterly Table 3 (2006–2025)"
        )
    else:
        gdp_notes = (
            f"Target panel: {target_q}. Reaches {panel.panel_start_year}Q1: {gdp_back}. "
            "Real GDP in millions of chained 2017 dollars divided by annual population "
            "(thousands × 1000 from FRED POP). Annual population is merged to all "
            "quarters in the same year; latest available population is forward-filled "
            "within each state when missing. FRED {ABBR}RQGSP typically begins "
            "2005Q1; earlier quarters are not fabricated."
        )
        gdp_source = "U.S. Bureau of Economic Analysis via FRED ({ABBR}RQGSP, {ABBR}POP)"
        home_notes = (
            f"Target panel: {target_q}. Reaches {panel.panel_start_year}Q1: {home_back}. "
            f"Parsed from local file {HOMEOWNERSHIP_FILE.name}; includes all years "
            "present in the Census file (earliest available ~2006). 2026Q1 excluded. "
            "margin_of_error: Census HVS Table 3 published margin of error "
            "(percentage points, 90% confidence level)."
        )
        home_source = "U.S. Census Bureau Housing Vacancy Survey, Table 3"

    notes = [
        {
            "variable": "real_gdp_per_capita",
            "source": gdp_source,
            "frequency": "quarterly",
            "seasonal_adjustment": "seasonally adjusted annual rate (real GDP)",
            "units": "chained 2017 dollars per person",
            "date_range": gdp_range,
            "output_file": "real_gdp_per_capita_by_state.csv",
            "notes": gdp_notes,
        },
        {
            "variable": "personal_income_per_capita",
            "source": "U.S. Bureau of Economic Analysis via FRED ({ABBR}OPCI or {ABBR}OTOT/POP)",
            "frequency": "quarterly",
            "seasonal_adjustment": "seasonally adjusted annual rate",
            "units": "current dollars per person",
            "date_range": pi_range,
            "output_file": "personal_income_per_capita_by_state.csv",
            "notes": (
                f"Target panel: {target_q}. Reaches {panel.panel_start_year}Q1: {pi_back}. "
                f"{pi_method} Annual population merged to quarters; forward-filled within "
                "state when missing."
            ),
        },
        {
            "variable": "monthly_unemployment_rate",
            "source": "U.S. Bureau of Labor Statistics LAUS via FRED ({ABBR}UR)",
            "frequency": "monthly",
            "seasonal_adjustment": "seasonally adjusted",
            "units": "percent",
            "date_range": ur_m_range,
            "output_file": "monthly_unemployment_rate_by_state.csv",
            "notes": (
                f"Target panel: {target_m}. Reaches {panel.panel_start_year}M1: {ur_back}. "
                "Statewide unemployment rate via Scrapling/FRED CSV ({ABBR}UR). "
                "2025M10 was blank in the FRED download for all 27 states; values were "
                "linearly interpolated from adjacent months (September and November 2025). "
                "2026 observations excluded. margin_of_error: 90% confidence interval "
                "half-width from BLS lanrderr.xlsx (latest LAUS release), scaled to each "
                "month's rate as rate × (ref_moe / ref_rate); units are percentage points."
            ),
        },
        {
            "variable": "quarterly_unemployment_rate",
            "source": "U.S. Bureau of Labor Statistics LAUS via FRED ({ABBR}UR)",
            "frequency": "quarterly",
            "seasonal_adjustment": "seasonally adjusted (average of monthly SA rates)",
            "units": "percent",
            "date_range": ur_q_range,
            "output_file": "unemployment_rate_by_state.csv",
            "notes": (
                f"Target panel: {target_q}. Reaches {panel.panel_start_year}Q1: {ur_back}. "
                "Arithmetic mean of three monthly unemployment rates; complete quarters "
                "only. Includes 2025Q4 using interpolated 2025M10 (see monthly notes). "
                "margin_of_error: mean of the three monthly margins of error in each "
                "quarter (percentage points)."
            ),
        },
        {
            "variable": "homeownership_rate",
            "source": home_source,
            "frequency": "quarterly",
            "seasonal_adjustment": "not seasonally adjusted",
            "units": "percent of occupied housing units owner-occupied",
            "date_range": home_range,
            "output_file": "homeownership_rate_by_state.csv",
            "notes": home_notes,
        },
    ]
    return pd.DataFrame(notes)


def assess_file(
    path: Path,
    value_col: str,
    is_monthly: bool = False,
) -> dict:
    df = pd.read_csv(path)
    states = sorted(df["state_abbr"].unique())
    target = set(TARGET_ABBRS)
    present = set(states)
    missing_states = sorted(target - present)
    extra_states = sorted(present - target)
    dups = int(df.duplicated(subset=["state_abbr", "date"]).sum())
    missing_vals = int(df[value_col].isna().sum()) if value_col in df.columns else 0

    return {
        "file": path.name,
        "rows": len(df),
        "unique_states": len(states),
        "min_date": df["date"].min(),
        "max_date": df["date"].max(),
        "missing_values": missing_vals,
        "duplicate_state_date_rows": dups,
        "all_27_states_present": present == target,
        "missing_states": ",".join(missing_states) if missing_states else "",
        "extra_states": ",".join(extra_states) if extra_states else "",
        "_state_list": states,
        "_is_monthly": is_monthly,
    }


def print_quality_report(reports: list[dict]) -> None:
    print(f"\n{'=' * 65}")
    print("QUALITY CHECKS")
    print(f"{'=' * 65}")
    for r in reports:
        print(f"\n--- {r['file']} ---")
        print(f"  Row count          : {r['rows']}")
        print(f"  Unique states      : {r['unique_states']}")
        print(f"  State list         : {r['_state_list']}")
        print(f"  All 27 present     : {r['all_27_states_present']}")
        print(f"  Date range         : {r['min_date']}  →  {r['max_date']}")
        print(f"  Missing values     : {r['missing_values']}")
        print(f"  Duplicate rows     : {r['duplicate_state_date_rows']}")
        if r["missing_states"]:
            print(f"  Missing states     : {r['missing_states']}")
        if r["extra_states"]:
            print(f"  Extra states       : {r['extra_states']}")


# ── Cleaned output (read output/, write cleaned_output/) ─────────────────────

SOURCE_NOTES_COLUMNS = [
    "variable",
    "source",
    "frequency",
    "seasonal_adjustment",
    "units",
    "date_range",
    "output_file",
    "notes",
]

def get_data_file_specs(panel: PanelConfig) -> dict[str, dict]:
    return {
        "real_gdp_per_capita_by_state.csv": {
            "columns": [
                "state",
                "state_abbr",
                "year",
                "quarter",
                "date",
                "real_gdp_per_capita",
                "real_gdp_millions_chained_2017_dollars",
                "population",
            ],
            "optional_columns": [
                "real_gdp_millions_chained_2017_dollars",
                "population",
            ],
            "value_col": "real_gdp_per_capita",
            "numeric_cols": [
                "real_gdp_per_capita",
                "real_gdp_millions_chained_2017_dollars",
                "population",
            ],
            "is_monthly": False,
            "expected_rows": panel.quarterly_expected_rows,
            "sort_cols": ["state_abbr", "year", "quarter"],
        },
        "personal_income_per_capita_by_state.csv": {
            "columns": [
                "state",
                "state_abbr",
                "year",
                "quarter",
                "date",
                "personal_income_per_capita",
            ],
            "optional_columns": [],
            "value_col": "personal_income_per_capita",
            "numeric_cols": ["personal_income_per_capita"],
            "is_monthly": False,
            "expected_rows": panel.quarterly_expected_rows,
            "sort_cols": ["state_abbr", "year", "quarter"],
        },
        "monthly_unemployment_rate_by_state.csv": {
            "columns": [
                "state",
                "state_abbr",
                "year",
                "month",
                "date",
                "unemployment_rate",
                "margin_of_error",
            ],
            "optional_columns": [],
            "value_col": "unemployment_rate",
            "numeric_cols": ["unemployment_rate", "margin_of_error"],
            "is_monthly": True,
            "expected_rows": panel.monthly_expected_rows,
            "sort_cols": ["state_abbr", "year", "month"],
        },
        "unemployment_rate_by_state.csv": {
            "columns": [
                "state",
                "state_abbr",
                "year",
                "quarter",
                "date",
                "unemployment_rate",
                "margin_of_error",
            ],
            "optional_columns": [],
            "value_col": "unemployment_rate",
            "numeric_cols": ["unemployment_rate", "margin_of_error"],
            "is_monthly": False,
            "expected_rows": panel.quarterly_expected_rows,
            "sort_cols": ["state_abbr", "year", "quarter"],
        },
        "homeownership_rate_by_state.csv": {
            "columns": [
                "state",
                "state_abbr",
                "year",
                "quarter",
                "date",
                "homeownership_rate",
                "margin_of_error",
            ],
            "optional_columns": [],
            "value_col": "homeownership_rate",
            "numeric_cols": ["homeownership_rate", "margin_of_error"],
            "is_monthly": False,
            "expected_rows": panel.quarterly_expected_rows,
            "sort_cols": ["state_abbr", "year", "quarter"],
        },
    }


DATA_FILE_SPECS = get_data_file_specs(PANEL_2010)


def _standardize_column_names(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    out.columns = (
        out.columns.astype(str)
        .str.strip()
        .str.lower()
        .str.replace(r"\s+", "_", regex=True)
    )
    drop_cols = [c for c in out.columns if not c or c.startswith("unnamed")]
    if drop_cols:
        out = out.drop(columns=drop_cols)
    return out


def _strip_string_columns(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    for col in out.select_dtypes(include=["object", "string"]).columns:
        out[col] = out[col].astype(str).str.strip()
        out[col] = out[col].replace({"nan": pd.NA, "None": pd.NA, "": pd.NA})
    return out


def _normalize_state_fields(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    if "state_abbr" not in out.columns:
        raise ValueError("Missing required column: state_abbr")
    out["state_abbr"] = out["state_abbr"].astype(str).str.strip().str.upper()
    out["state"] = out["state_abbr"].map(STATE_MAP)
    unknown = sorted(out.loc[out["state"].isna(), "state_abbr"].unique())
    if unknown:
        raise ValueError(f"Unknown state abbreviations: {unknown}")
    return out


def _ensure_iso_dates(df: pd.DataFrame, is_monthly: bool) -> pd.DataFrame:
    out = df.copy()
    if "date" in out.columns:
        out["date"] = pd.to_datetime(out["date"], errors="coerce").dt.strftime("%Y-%m-%d")
    elif is_monthly:
        out["date"] = out.apply(
            lambda r: month_end_date(int(r["year"]), int(r["month"])), axis=1
        )
    else:
        out["date"] = out.apply(
            lambda r: quarter_end_date(int(r["year"]), int(r["quarter"])), axis=1
        )
    return out


def _clean_dataframe(df: pd.DataFrame, spec: dict, panel: PanelConfig) -> pd.DataFrame:
    out = _standardize_column_names(df)
    out = _strip_string_columns(out)
    out = _normalize_state_fields(out)

    out = out[out["state_abbr"].isin(TARGET_ABBRS)].copy()

    out["year"] = pd.to_numeric(out["year"], errors="coerce").astype("Int64")
    if spec["is_monthly"]:
        out["month"] = pd.to_numeric(out["month"], errors="coerce").astype("Int64")
        out = out.dropna(subset=["year", "month"])
        out["year"] = out["year"].astype(int)
        out["month"] = out["month"].astype(int)
        out = filter_monthly_panel(out, panel)
    else:
        out["quarter"] = pd.to_numeric(out["quarter"], errors="coerce").astype("Int64")
        out = out.dropna(subset=["year", "quarter"])
        out["year"] = out["year"].astype(int)
        out["quarter"] = out["quarter"].astype(int)
        out = filter_quarterly_panel(out, panel)

    out = _ensure_iso_dates(out, is_monthly=spec["is_monthly"])

    for col in spec.get("numeric_cols", []):
        if col in out.columns:
            out[col] = pd.to_numeric(out[col], errors="coerce")

    out = (
        out.sort_values(spec["sort_cols"])
        .drop_duplicates(subset=["state_abbr", "date"], keep="last")
        .reset_index(drop=True)
    )

    missing_optional = [c for c in spec.get("optional_columns", []) if c not in out.columns]
    if missing_optional:
        log(f"  WARNING: missing optional columns (omitted): {', '.join(missing_optional)}")

    allowed_cols = set(spec["columns"]) | set(spec.get("optional_columns", []))
    drop_extra = [c for c in out.columns if c not in allowed_cols]
    if drop_extra:
        out = out.drop(columns=drop_extra)

    final_cols = [c for c in spec["columns"] if c in out.columns]
    return out[final_cols]


def _clean_source_notes(df: pd.DataFrame) -> pd.DataFrame:
    out = _standardize_column_names(df)
    out = _strip_string_columns(out)
    final_cols = [c for c in SOURCE_NOTES_COLUMNS if c in out.columns]
    extra_cols = [c for c in out.columns if c not in final_cols]
    return out[final_cols + extra_cols]


def _assess_cleaned_file(path: Path, spec: dict, panel: PanelConfig) -> dict:
    df = pd.read_csv(path)
    value_col = spec["value_col"]
    states = sorted(df["state_abbr"].unique())
    target = set(TARGET_ABBRS)
    present = set(states)
    missing_states = sorted(target - present)
    extra_states = sorted(present - target)
    dups = int(df.duplicated(subset=["state_abbr", "date"]).sum())
    missing_vals = int(df[value_col].isna().sum()) if value_col in df.columns else 0
    rows = len(df)
    expected_rows = spec["expected_rows"]
    is_monthly = spec["is_monthly"]

    all_27 = present == target
    no_dups = dups == 0
    no_extra = not extra_states
    row_count_matches = rows == expected_rows

    min_date = df["date"].min()
    max_date = df["date"].max()
    target_start = _panel_target_start_date(panel, is_monthly)
    partial_date_range = pd.to_datetime(min_date) > pd.to_datetime(target_start)

    notes: list[str] = []
    if partial_date_range:
        notes.append(
            f"Earliest date {min_date} is after target start {target_start}; "
            "source does not reach 2001."
        )
    if not row_count_matches:
        notes.append(f"Row count {rows} differs from expected {expected_rows}.")
    if missing_vals:
        notes.append(f"{missing_vals} missing value(s) in {value_col}.")
    if missing_states:
        notes.append(f"Missing states: {', '.join(missing_states)}.")
    if extra_states:
        notes.append(f"Extra states: {', '.join(extra_states)}.")
    if dups:
        notes.append(f"{dups} duplicate state-date row(s).")

    if not all_27 or dups or extra_states:
        status = "error"
    elif partial_date_range:
        status = "partial_date_range"
    elif missing_vals > 0 or not row_count_matches:
        status = "warning"
    else:
        status = "ok"

    return {
        "file": path.name,
        "rows": rows,
        "unique_states": len(states),
        "min_date": min_date,
        "max_date": max_date,
        "missing_values": missing_vals,
        "duplicate_state_date_rows": dups,
        "all_27_states_present": all_27,
        "missing_states": ",".join(missing_states) if missing_states else "",
        "extra_states": ",".join(extra_states) if extra_states else "",
        "expected_rows": expected_rows,
        "row_count_matches_expected": row_count_matches,
        "status": status,
        "notes": " ".join(notes),
    }


def run_cleanup_step(
    input_dir: Path | None = None,
    output_dir: Path | None = None,
    panel: PanelConfig | None = None,
) -> list[dict]:
    """
    Read assembled CSVs from input_dir, clean/standardize, write to output_dir.
    Never modifies files in the original input folder.
    """
    panel = panel or PANEL_2010
    input_dir = input_dir or panel.output_dir
    output_dir = output_dir or panel.cleaned_dir
    output_dir.mkdir(parents=True, exist_ok=True)
    specs = get_data_file_specs(panel)

    log(f"\n{'=' * 65}")
    log("CLEANUP: standardizing CSVs for Stata/R import")
    log(f"Panel         : {panel.label}")
    log(f"Reading from  : {input_dir}")
    log(f"Writing to    : {output_dir}")
    log(f"{'=' * 65}")

    cleaned_reports: list[dict] = []

    for filename, spec in specs.items():
        src = input_dir / filename
        if not src.exists():
            raise FileNotFoundError(f"Expected input file not found: {src}")
        raw = pd.read_csv(src)
        cleaned = _clean_dataframe(raw, spec, panel)
        dest = output_dir / filename
        cleaned.to_csv(dest, index=False, na_rep="")
        report = _assess_cleaned_file(dest, spec, panel)
        cleaned_reports.append(report)
        log(f"  Cleaned: {filename} ({len(cleaned)} rows, {report['min_date']} → {report['max_date']})")

    src_notes = input_dir / "source_notes.csv"
    if not src_notes.exists():
        raise FileNotFoundError(f"Expected input file not found: {src_notes}")
    notes_cleaned = _clean_source_notes(pd.read_csv(src_notes))
    notes_cleaned.to_csv(output_dir / "source_notes.csv", index=False, na_rep="")

    report_df = pd.DataFrame(cleaned_reports)
    report_df.to_csv(output_dir / "data_quality_report.csv", index=False, na_rep="")

    log(f"\n{'=' * 65}")
    log(f"Cleaned files written to {output_dir.name}/")
    log(f"{'=' * 65}")
    for r in cleaned_reports:
        log(f"\n  {r['file']}")
        log(f"    rows            : {r['rows']} (expected {r['expected_rows']})")
        log(f"    unique states   : {r['unique_states']}")
        log(f"    date range      : {r['min_date']} → {r['max_date']}")
        log(f"    missing values  : {r['missing_values']}")
        log(f"    duplicates      : {r['duplicate_state_date_rows']}")
        log(f"    status          : {r['status']}")
        if r.get("notes"):
            log(f"    notes           : {r['notes']}")
    return cleaned_reports


def run_panel_build(panel: PanelConfig) -> list[dict]:
    global FAILED_SERIES
    FAILED_SERIES = []
    unemployment_failed: list[str] = []

    panel.output_dir.mkdir(parents=True, exist_ok=True)
    panel.fred_cache_dir.mkdir(parents=True, exist_ok=True)

    log("=" * 65)
    log(f"Building state economic datasets — {panel.label}")
    log(f"Output directory : {panel.output_dir}")
    log(f"Cleaned output   : {panel.cleaned_dir}")
    log(f"Panel range      : {panel.panel_start_year} – {panel.panel_end_year}")
    log(f"FRED start date  : {panel.start_date}")
    log(f"FRED cache       : {panel.fred_cache_dir}")
    log(f"HTTP client      : Scrapling FetcherSession (curl_cffi / Chrome)")
    log("=" * 65)

    with FetcherSession(
        impersonate="chrome",
        stealthy_headers=True,
        timeout=20,
        retries=2,
        retry_delay=2,
    ) as fred_session:
        pop_raw, pop_failed = fetch_state_series(
            fred_session, "POP", "Annual population by state (FRED {ABBR}POP)", panel
        )
        pop_annual = population_from_fred(pop_raw)
        if pop_failed:
            log(f"  WARNING: POP failed series: {pop_failed}")

        gdp_out = build_real_gdp_per_capita(fred_session, pop_annual, panel)
        pi_out = build_personal_income_per_capita(fred_session, pop_annual, panel)
        monthly_out, quarterly_out, unemployment_failed = build_unemployment(
            fred_session, panel
        )
        home_out = build_homeownership(panel, session=fred_session)

    data_outputs = {
        "real_gdp_per_capita_by_state.csv": gdp_out,
        "personal_income_per_capita_by_state.csv": pi_out,
        "monthly_unemployment_rate_by_state.csv": monthly_out,
        "unemployment_rate_by_state.csv": quarterly_out,
        "homeownership_rate_by_state.csv": home_out,
    }

    source_notes = build_source_notes(panel, data_outputs)

    outputs = {**data_outputs, "source_notes.csv": source_notes}

    log(f"\n{'=' * 65}")
    log(f"Saving CSV files ({panel.label})")
    log(f"{'=' * 65}")
    for name, df in outputs.items():
        path = panel.output_dir / name
        df.to_csv(path, index=False)
        if name.endswith(".csv") and name != "source_notes.csv":
            if "month" in df.columns:
                log(f"  Saved: {name} ({len(df)} rows, {_monthly_range_label(df)})")
            else:
                log(f"  Saved: {name} ({len(df)} rows, {_quarterly_range_label(df)})")
        else:
            log(f"  Saved: {path}")

    value_cols = {
        "real_gdp_per_capita_by_state.csv": "real_gdp_per_capita",
        "personal_income_per_capita_by_state.csv": "personal_income_per_capita",
        "monthly_unemployment_rate_by_state.csv": "unemployment_rate",
        "unemployment_rate_by_state.csv": "unemployment_rate",
        "homeownership_rate_by_state.csv": "homeownership_rate",
    }
    reports = [
        assess_file(
            panel.output_dir / name,
            value_cols[name],
            is_monthly=(name == "monthly_unemployment_rate_by_state.csv"),
        )
        for name in value_cols
    ]
    report_df = pd.DataFrame([
        {k: v for k, v in r.items() if not k.startswith("_")}
        for r in reports
    ])
    report_path = panel.output_dir / "data_quality_report.csv"
    report_df.to_csv(report_path, index=False)
    log(f"  Saved: {report_path}")

    print_quality_report(reports)

    if FAILED_SERIES:
        log(f"  ⚠ Failed FRED series ({len(FAILED_SERIES)}): {FAILED_SERIES}")
    if unemployment_failed:
        log(f"  ⚠ Unemployment download failures: {unemployment_failed}")

    cleaned_reports = run_cleanup_step(
        input_dir=panel.output_dir,
        output_dir=panel.cleaned_dir,
        panel=panel,
    )

    log(f"\n{'=' * 65}")
    log(f"PANEL SUMMARY — {panel.label}")
    log(f"{'=' * 65}")
    for r in cleaned_reports:
        reaches = "complete from panel start" if r["status"] == "ok" else r["status"]
        log(
            f"  {r['file']}: {r['min_date']} → {r['max_date']} "
            f"({r['rows']} rows) — {reaches}"
        )

    return cleaned_reports


def main() -> None:
    parser = argparse.ArgumentParser(description="Build state economic datasets")
    parser.add_argument(
        "--extended",
        action="store_true",
        help="Build only the 2001–2025 extended panel (cleaned_output_2001_2025/)",
    )
    parser.add_argument(
        "--standard",
        action="store_true",
        help="Build only the 2010–2025 standard panel (cleaned_output/)",
    )
    parser.add_argument(
        "--full-panel",
        action="store_true",
        help=(
            "Build 2001–2025 full panel with annual RGSP/Table 13 backfill "
            "(cleaned_output_2001_2025_full_panel/)"
        ),
    )
    args = parser.parse_args()

    flags = sum([args.extended, args.standard, args.full_panel])
    if flags > 1:
        raise SystemExit("Use only one of --extended, --standard, or --full-panel.")

    if args.full_panel:
        panels = [PANEL_2001_FULL]
    elif args.extended:
        panels = [PANEL_2001]
    elif args.standard:
        panels = [PANEL_2010]
    else:
        panels = [PANEL_2010, PANEL_2001]

    all_reports: dict[str, list[dict]] = {}
    for panel in panels:
        all_reports[panel.label] = run_panel_build(panel)

    if len(panels) == 2:
        log(f"\n{'=' * 65}")
        log("2001 vs 2010 COMPLETENESS")
        log(f"{'=' * 65}")
        for r in all_reports["2001-2025"]:
            back_to_2001 = "yes" if r["status"] == "ok" else "no"
            log(f"  {r['file']}: back to 2001 = {back_to_2001} ({r['status']})")

    log("\nDone.")


if __name__ == "__main__":
    main()
